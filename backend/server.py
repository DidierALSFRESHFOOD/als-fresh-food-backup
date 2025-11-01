from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, Request, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'als-groupe-frigo-kpi-secret-key-2025')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24 * 7  # 7 days

# Emergent Auth Configuration
EMERGENT_SESSION_API = 'https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data'

# Create the main app
app = FastAPI(title="ALS GROUPE FRIGO KPI API")
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    password_hash: Optional[str] = None
    picture: Optional[str] = None
    role: str  # Admin_Directeur, Assistante_Direction, Directrice_Clientele, etc.
    division: Optional[str] = None  # ALS FRESH FOOD / ALS PHARMA
    region: Optional[str] = None  # IDF / HDF
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TranslationKey(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    key: str
    value: str
    lang: str = "fr-FR"
    updated_by: str
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Compte(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    raison_sociale: str
    division: str  # ALS FRESH FOOD / ALS PHARMA
    adresse: Optional[str] = None
    ville: Optional[str] = None
    code_postal: Optional[str] = None
    region: str  # IDF / HDF
    secteur: Optional[str] = None
    taille: Optional[str] = None  # TPE / PME / enseigne / groupe
    contact_nom: Optional[str] = None
    contact_poste: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    contact_telephone: Optional[str] = None
    source: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CompteCreate(BaseModel):
    raison_sociale: str
    division: str
    adresse: Optional[str] = None
    ville: Optional[str] = None
    code_postal: Optional[str] = None
    region: str
    secteur: Optional[str] = None
    taille: Optional[str] = None
    contact_nom: Optional[str] = None
    contact_poste: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    contact_telephone: Optional[str] = None
    source: Optional[str] = None

class Opportunite(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    compte_id: str
    type_besoin: Optional[str] = None
    volumes_estimes: Optional[str] = None
    temperatures: Optional[str] = None
    frequence: Optional[str] = None
    marchandises: Optional[str] = None
    depart: Optional[str] = None
    arrivee: Optional[str] = None
    contraintes_horaires: Optional[str] = None
    urgence: Optional[str] = None
    commercial_responsable: str
    date_premier_contact: Optional[datetime] = None
    canal: Optional[str] = None
    statut: str = "Prospecté"  # Prospecté, En discussion, Devis envoyé, Négociation, Signé, Perdu
    montant_estime: Optional[float] = None
    prochaine_relance: Optional[datetime] = None
    commentaires: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QualityRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    compte_id: str
    division: str
    region: str
    periode: str
    type_prestation: Optional[str] = None
    taux_service: Optional[float] = None
    nb_incidents: int = 0
    score_satisfaction: Optional[float] = None
    commentaires: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Incident(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    quality_record_id: str
    type: str
    gravite: str  # Faible / Moyen / Critique
    description: str
    statut: str = "Ouvert"
    action_corrective: Optional[str] = None
    closed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SurveyResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    compte_id: str
    division: str
    periode: str
    note_globale: Optional[int] = None  # 0-9
    commentaires: Optional[str] = None
    submitted_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SurveyScore(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    response_id: str
    theme: str  # Conducteurs, Matériel, Tournées, etc.
    item_key: str
    score: int  # 0-9

# ==================== AUTH MODELS ====================

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class RegisterRequest(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "DevCo_IDF"

class SessionRequest(BaseModel):
    session_id: str

class TokenResponse(BaseModel):
    token: str
    user: dict

# ==================== AUTH HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_jwt_token(user_id: str) -> str:
    payload = {
        'user_id': user_id,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS),
        'iat': datetime.now(timezone.utc)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> User:
    # Try cookie first
    session_token = request.cookies.get('session_token')
    
    # Try Authorization header as fallback
    if not session_token:
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            session_token = auth_header.replace('Bearer ', '')
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Non authentifié")
    
    # Try to find user session
    session_doc = await db.user_sessions.find_one({
        'session_token': session_token,
        'expires_at': {'$gt': datetime.now(timezone.utc).isoformat()}
    })
    
    if session_doc:
        user_doc = await db.users.find_one({'id': session_doc['user_id']}, {'_id': 0})
        if user_doc:
            return User(**user_doc)
    
    # Try JWT token
    try:
        payload = jwt.decode(session_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_doc = await db.users.find_one({'id': payload['user_id']}, {'_id': 0})
        if user_doc:
            return User(**user_doc)
    except:
        pass
    
    raise HTTPException(status_code=401, detail="Session invalide")

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(data: RegisterRequest):
    # Check if user exists
    existing = await db.users.find_one({'email': data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    
    user = User(
        email=data.email,
        name=data.name,
        password_hash=hash_password(data.password),
        role=data.role
    )
    
    user_dict = user.model_dump()
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    await db.users.insert_one(user_dict)
    
    token = create_jwt_token(user.id)
    return TokenResponse(token=token, user=user.model_dump(exclude={'password_hash'}))

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: LoginRequest):
    user_doc = await db.users.find_one({'email': data.email}, {'_id': 0})
    if not user_doc or not user_doc.get('password_hash'):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    if not verify_password(data.password, user_doc['password_hash']):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    user = User(**user_doc)
    token = create_jwt_token(user.id)
    return TokenResponse(token=token, user=user.model_dump(exclude={'password_hash'}))

@api_router.post("/auth/google-session")
async def google_session(data: SessionRequest, response: Response):
    # Call Emergent Auth API
    try:
        resp = requests.get(
            EMERGENT_SESSION_API,
            headers={'X-Session-ID': data.session_id},
            timeout=10
        )
        resp.raise_for_status()
        session_data = resp.json()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur authentification Google: {str(e)}")
    
    # Check if user exists
    user_doc = await db.users.find_one({'email': session_data['email']}, {'_id': 0})
    
    if not user_doc:
        # Create new user
        user = User(
            email=session_data['email'],
            name=session_data['name'],
            picture=session_data.get('picture'),
            role='DevCo_IDF'  # Default role
        )
        user_dict = user.model_dump()
        user_dict['created_at'] = user_dict['created_at'].isoformat()
        await db.users.insert_one(user_dict)
    else:
        user = User(**user_doc)
    
    # Create session
    session = UserSession(
        user_id=user.id,
        session_token=session_data['session_token'],
        expires_at=datetime.now(timezone.utc) + timedelta(days=7)
    )
    
    session_dict = session.model_dump()
    session_dict['expires_at'] = session_dict['expires_at'].isoformat()
    session_dict['created_at'] = session_dict['created_at'].isoformat()
    await db.user_sessions.insert_one(session_dict)
    
    # Set cookie
    response.set_cookie(
        key='session_token',
        value=session_data['session_token'],
        httponly=True,
        secure=True,
        samesite='none',
        max_age=7 * 24 * 60 * 60,
        path='/'
    )
    
    return {'user': user.model_dump(exclude={'password_hash'}), 'session_token': session_data['session_token']}

@api_router.get("/auth/me")
async def get_me(user: User = Depends(get_current_user)):
    return user.model_dump(exclude={'password_hash'})

@api_router.post("/auth/logout")
async def logout(response: Response, request: Request):
    session_token = request.cookies.get('session_token')
    if session_token:
        await db.user_sessions.delete_many({'session_token': session_token})
    response.delete_cookie('session_token', path='/')
    return {'message': 'Déconnexion réussie'}

# ==================== COMPTES ROUTES ====================

@api_router.post("/comptes", response_model=Compte)
async def create_compte(data: CompteCreate, user: User = Depends(get_current_user)):
    compte = Compte(**data.model_dump(), created_by=user.id)
    compte_dict = compte.model_dump()
    compte_dict['created_at'] = compte_dict['created_at'].isoformat()
    await db.comptes.insert_one(compte_dict)
    return compte

@api_router.get("/comptes", response_model=List[Compte])
async def get_comptes(user: User = Depends(get_current_user)):
    query = {}
    if user.region and user.role not in ['Admin_Directeur', 'Assistante_Direction']:
        query['region'] = user.region
    
    comptes = await db.comptes.find(query, {'_id': 0}).to_list(1000)
    for c in comptes:
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return comptes

@api_router.get("/comptes/{compte_id}", response_model=Compte)
async def get_compte(compte_id: str, user: User = Depends(get_current_user)):
    compte = await db.comptes.find_one({'id': compte_id}, {'_id': 0})
    if not compte:
        raise HTTPException(status_code=404, detail="Compte non trouvé")
    if isinstance(compte.get('created_at'), str):
        compte['created_at'] = datetime.fromisoformat(compte['created_at'])
    return Compte(**compte)

@api_router.delete("/comptes/{compte_id}")
async def delete_compte(compte_id: str, user: User = Depends(get_current_user)):
    result = await db.comptes.delete_one({'id': compte_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Compte non trouvé")
    
    # Also delete related opportunites
    await db.opportunites.delete_many({'compte_id': compte_id})
    await db.quality_records.delete_many({'compte_id': compte_id})
    
    return {'message': 'Compte et données associées supprimés'}

# ==================== OPPORTUNITES ROUTES ====================

@api_router.post("/opportunites", response_model=Opportunite)
async def create_opportunite(data: Opportunite, user: User = Depends(get_current_user)):
    opp = Opportunite(**data.model_dump(), commercial_responsable=user.id)
    opp_dict = opp.model_dump()
    opp_dict['created_at'] = opp_dict['created_at'].isoformat()
    if opp_dict.get('date_premier_contact'):
        opp_dict['date_premier_contact'] = opp_dict['date_premier_contact'].isoformat()
    if opp_dict.get('prochaine_relance'):
        opp_dict['prochaine_relance'] = opp_dict['prochaine_relance'].isoformat()
    await db.opportunites.insert_one(opp_dict)
    return opp

@api_router.get("/opportunites", response_model=List[Opportunite])
async def get_opportunites(user: User = Depends(get_current_user)):
    query = {}
    if user.role in ['DevCo_IDF', 'DevCo_HDF']:
        query['commercial_responsable'] = user.id
    
    opps = await db.opportunites.find(query, {'_id': 0}).to_list(1000)
    for o in opps:
        if isinstance(o.get('created_at'), str):
            o['created_at'] = datetime.fromisoformat(o['created_at'])
        if o.get('date_premier_contact') and isinstance(o['date_premier_contact'], str):
            o['date_premier_contact'] = datetime.fromisoformat(o['date_premier_contact'])
        if o.get('prochaine_relance') and isinstance(o['prochaine_relance'], str):
            o['prochaine_relance'] = datetime.fromisoformat(o['prochaine_relance'])
    return opps

@api_router.delete("/opportunites/{opp_id}")
async def delete_opportunite(opp_id: str, user: User = Depends(get_current_user)):
    result = await db.opportunites.delete_one({'id': opp_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Opportunité non trouvée")
    return {'message': 'Opportunité supprimée'}

# ==================== QUALITY ROUTES ====================

@api_router.post("/quality", response_model=QualityRecord)
async def create_quality_record(data: QualityRecord, user: User = Depends(get_current_user)):
    record_dict = data.model_dump()
    record_dict['created_at'] = record_dict['created_at'].isoformat()
    await db.quality_records.insert_one(record_dict)
    return data

@api_router.get("/quality", response_model=List[QualityRecord])
async def get_quality_records(user: User = Depends(get_current_user)):
    records = await db.quality_records.find({}, {'_id': 0}).to_list(1000)
    for r in records:
        if isinstance(r.get('created_at'), str):
            r['created_at'] = datetime.fromisoformat(r['created_at'])
    return records

@api_router.delete("/quality/{quality_id}")
async def delete_quality_record(quality_id: str, user: User = Depends(get_current_user)):
    result = await db.quality_records.delete_one({'id': quality_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fiche qualité non trouvée")
    
    # Also delete related incidents
    await db.incidents.delete_many({'quality_record_id': quality_id})
    
    return {'message': 'Fiche qualité et incidents associés supprimés'}

@api_router.post("/incidents", response_model=Incident)
async def create_incident(data: Incident, user: User = Depends(get_current_user)):
    incident_dict = data.model_dump()
    incident_dict['created_at'] = incident_dict['created_at'].isoformat()
    if incident_dict.get('closed_at'):
        incident_dict['closed_at'] = incident_dict['closed_at'].isoformat()
    await db.incidents.insert_one(incident_dict)
    return data

@api_router.get("/incidents", response_model=List[Incident])
async def get_incidents(user: User = Depends(get_current_user)):
    incidents = await db.incidents.find({}, {'_id': 0}).to_list(1000)
    for i in incidents:
        if isinstance(i.get('created_at'), str):
            i['created_at'] = datetime.fromisoformat(i['created_at'])
        if i.get('closed_at') and isinstance(i['closed_at'], str):
            i['closed_at'] = datetime.fromisoformat(i['closed_at'])
    return incidents

@api_router.delete("/incidents/{incident_id}")
async def delete_incident(incident_id: str, user: User = Depends(get_current_user)):
    result = await db.incidents.delete_one({'id': incident_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Incident non trouvé")
    return {'message': 'Incident supprimé'}

# ==================== SURVEY ROUTES ====================

@api_router.post("/surveys/responses", response_model=SurveyResponse)
async def create_survey_response(data: SurveyResponse):
    response_dict = data.model_dump()
    response_dict['submitted_at'] = response_dict['submitted_at'].isoformat()
    await db.survey_responses.insert_one(response_dict)
    return data

@api_router.post("/surveys/scores", response_model=SurveyScore)
async def create_survey_score(data: SurveyScore):
    await db.survey_scores.insert_one(data.model_dump())
    return data

@api_router.get("/surveys/responses", response_model=List[SurveyResponse])
async def get_survey_responses(user: User = Depends(get_current_user)):
    responses = await db.survey_responses.find({}, {'_id': 0}).to_list(1000)
    for r in responses:
        if isinstance(r.get('submitted_at'), str):
            r['submitted_at'] = datetime.fromisoformat(r['submitted_at'])
    return responses

# ==================== DASHBOARD ROUTES ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: User = Depends(get_current_user)):
    # Commercial Stats
    total_comptes = await db.comptes.count_documents({})
    total_opps = await db.opportunites.count_documents({})
    opps_signees = await db.opportunites.count_documents({'statut': 'Signé'})
    
    # Calculate CA signé
    pipeline = [
        {'$match': {'statut': 'Signé'}},
        {'$group': {'_id': None, 'total': {'$sum': '$montant_estime'}}}
    ]
    ca_result = await db.opportunites.aggregate(pipeline).to_list(1)
    ca_signe = ca_result[0]['total'] if ca_result and ca_result[0]['total'] else 0
    
    # Quality Stats
    total_quality = await db.quality_records.count_documents({})
    total_incidents = await db.incidents.count_documents({})
    incidents_ouverts = await db.incidents.count_documents({'statut': 'Ouvert'})
    
    # Average satisfaction
    satisfaction_pipeline = [
        {'$group': {'_id': None, 'avg': {'$avg': '$score_satisfaction'}}}
    ]
    satisfaction_result = await db.quality_records.aggregate(satisfaction_pipeline).to_list(1)
    avg_satisfaction = satisfaction_result[0]['avg'] if satisfaction_result and satisfaction_result[0]['avg'] else 0
    
    return {
        'commercial': {
            'total_comptes': total_comptes,
            'total_opportunites': total_opps,
            'opportunites_signees': opps_signees,
            'ca_signe': round(ca_signe, 2)
        },
        'qualite': {
            'total_quality_records': total_quality,
            'total_incidents': total_incidents,
            'incidents_ouverts': incidents_ouverts,
            'score_satisfaction_moyen': round(avg_satisfaction, 1) if avg_satisfaction else 0
        }
    }

# ==================== ADMIN ROUTES ====================

@api_router.get("/admin/users", response_model=List[User])
async def get_all_users(user: User = Depends(get_current_user)):
    if user.role != 'Admin_Directeur':
        raise HTTPException(status_code=403, detail="Accès réservé à la Direction commerciale")
    
    users = await db.users.find({}, {'_id': 0}).to_list(1000)
    for u in users:
        if isinstance(u.get('created_at'), str):
            u['created_at'] = datetime.fromisoformat(u['created_at'])
    return users

@api_router.post("/admin/users", response_model=User)
async def create_user_admin(data: User, user: User = Depends(get_current_user)):
    if user.role != 'Admin_Directeur':
        raise HTTPException(status_code=403, detail="Accès réservé à la Direction commerciale")
    
    existing = await db.users.find_one({'email': data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    
    new_user = User(**data.model_dump())
    user_dict = new_user.model_dump()
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    await db.users.insert_one(user_dict)
    return new_user

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, user: User = Depends(get_current_user)):
    if user.role != 'Admin_Directeur':
        raise HTTPException(status_code=403, detail="Accès réservé à la Direction commerciale")
    
    # Don't allow deleting yourself
    if user_id == user.id:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas supprimer votre propre compte")
    
    result = await db.users.delete_one({'id': user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Also delete user sessions
    await db.user_sessions.delete_many({'user_id': user_id})
    
    return {'message': 'Utilisateur supprimé avec succès'}

@api_router.get("/admin/translations/init")
async def init_translations(user: User = Depends(get_current_user)):
    if user.role != 'Admin_Directeur':
        raise HTTPException(status_code=403, detail="Accès réservé à la Direction commerciale")
    
    # Default translations for the app
    default_translations = [
        {'key': 'app.title', 'value': 'Suivi Activité Commerciale', 'lang': 'fr-FR'},
        {'key': 'app.subtitle', 'value': 'ALS FRESH FOOD • ALS PHARMA', 'lang': 'fr-FR'},
        {'key': 'login.title', 'value': 'Connexion', 'lang': 'fr-FR'},
        {'key': 'login.description', 'value': 'Accédez à votre espace commercial et qualité', 'lang': 'fr-FR'},
        {'key': 'login.email', 'value': 'Email', 'lang': 'fr-FR'},
        {'key': 'login.password', 'value': 'Mot de passe', 'lang': 'fr-FR'},
        {'key': 'login.submit', 'value': 'Se connecter', 'lang': 'fr-FR'},
        {'key': 'login.google', 'value': 'Continuer avec Google', 'lang': 'fr-FR'},
        {'key': 'nav.dashboard', 'value': 'Tableau de bord', 'lang': 'fr-FR'},
        {'key': 'nav.comptes', 'value': 'Clients / Prospects', 'lang': 'fr-FR'},
        {'key': 'nav.opportunites', 'value': 'Opportunités', 'lang': 'fr-FR'},
        {'key': 'nav.qualite', 'value': 'Qualité Service', 'lang': 'fr-FR'},
        {'key': 'nav.incidents', 'value': 'Incidents', 'lang': 'fr-FR'},
        {'key': 'nav.satisfaction', 'value': 'Satisfaction', 'lang': 'fr-FR'},
        {'key': 'nav.admin', 'value': 'Administration', 'lang': 'fr-FR'},
        {'key': 'nav.logout', 'value': 'Déconnexion', 'lang': 'fr-FR'},
        {'key': 'dashboard.welcome', 'value': 'Bienvenue', 'lang': 'fr-FR'},
        {'key': 'dashboard.commercial', 'value': 'Module Commercial', 'lang': 'fr-FR'},
        {'key': 'dashboard.qualite', 'value': 'Module Qualité & Service Clientèle', 'lang': 'fr-FR'},
        {'key': 'dashboard.ca_signe', 'value': 'CA Signé', 'lang': 'fr-FR'},
        {'key': 'dashboard.clients_prospects', 'value': 'Clients / Prospects', 'lang': 'fr-FR'},
        {'key': 'dashboard.opportunites', 'value': 'Opportunités', 'lang': 'fr-FR'},
        {'key': 'dashboard.opportunites_signees', 'value': 'Opportunités Signées', 'lang': 'fr-FR'},
        {'key': 'dashboard.score_satisfaction', 'value': 'Score Satisfaction', 'lang': 'fr-FR'},
        {'key': 'dashboard.fiches_qualite', 'value': 'Fiches Qualité', 'lang': 'fr-FR'},
        {'key': 'dashboard.incidents_total', 'value': 'Incidents Total', 'lang': 'fr-FR'},
        {'key': 'dashboard.incidents_ouverts', 'value': 'Incidents Ouverts', 'lang': 'fr-FR'},
        {'key': 'comptes.title', 'value': 'Clients / Prospects', 'lang': 'fr-FR'},
        {'key': 'comptes.description', 'value': 'Gestion de la base clients et prospects', 'lang': 'fr-FR'},
        {'key': 'comptes.add', 'value': 'Nouveau Compte', 'lang': 'fr-FR'},
        {'key': 'comptes.search', 'value': 'Rechercher par raison sociale, ville, contact...', 'lang': 'fr-FR'},
        {'key': 'opportunites.title', 'value': 'Opportunités Commerciales', 'lang': 'fr-FR'},
        {'key': 'opportunites.description', 'value': 'Pipeline de vente et suivi des opportunités', 'lang': 'fr-FR'},
        {'key': 'opportunites.add', 'value': 'Nouvelle Opportunité', 'lang': 'fr-FR'},
        {'key': 'qualite.title', 'value': 'Qualité de Service', 'lang': 'fr-FR'},
        {'key': 'qualite.description', 'value': 'Suivi de la satisfaction client et taux de service', 'lang': 'fr-FR'},
        {'key': 'qualite.add', 'value': 'Nouvelle Fiche Qualité', 'lang': 'fr-FR'},
        {'key': 'incidents.title', 'value': 'Incidents Qualité', 'lang': 'fr-FR'},
        {'key': 'incidents.description', 'value': 'Suivi et résolution des incidents', 'lang': 'fr-FR'},
        {'key': 'incidents.add', 'value': 'Nouvel Incident', 'lang': 'fr-FR'},
        {'key': 'admin.title', 'value': 'Administration', 'lang': 'fr-FR'},
        {'key': 'admin.description', 'value': 'Gestion des utilisateurs, textes et paramètres', 'lang': 'fr-FR'},
        {'key': 'admin.tab_translations', 'value': 'Textes & Libellés', 'lang': 'fr-FR'},
        {'key': 'admin.tab_users', 'value': 'Utilisateurs', 'lang': 'fr-FR'},
        {'key': 'admin.tab_settings', 'value': 'Paramètres', 'lang': 'fr-FR'},
        {'key': 'admin.add_translation', 'value': 'Nouvelle Clé', 'lang': 'fr-FR'},
        {'key': 'admin.add_user', 'value': 'Nouvel Utilisateur', 'lang': 'fr-FR'},
        {'key': 'common.save', 'value': 'Enregistrer', 'lang': 'fr-FR'},
        {'key': 'common.delete', 'value': 'Supprimer', 'lang': 'fr-FR'},
        {'key': 'common.cancel', 'value': 'Annuler', 'lang': 'fr-FR'},
        {'key': 'common.create', 'value': 'Créer', 'lang': 'fr-FR'},
        {'key': 'common.edit', 'value': 'Modifier', 'lang': 'fr-FR'},
    ]
    
    # Check if translations already exist
    existing_count = await db.translation_keys.count_documents({})
    if existing_count > 0:
        return {'message': 'Traductions déjà initialisées', 'count': existing_count}
    
    # Insert default translations
    for trans in default_translations:
        translation = TranslationKey(**trans, updated_by=user.id)
        trans_dict = translation.model_dump()
        trans_dict['updated_at'] = trans_dict['updated_at'].isoformat()
        await db.translation_keys.insert_one(trans_dict)
    
    return {'message': f'{len(default_translations)} traductions initialisées avec succès'}

@api_router.get("/admin/translations", response_model=List[TranslationKey])
async def get_translations(user: User = Depends(get_current_user)):
    if user.role != 'Admin_Directeur':
        raise HTTPException(status_code=403, detail="Accès réservé à la Direction commerciale")
    
    translations = await db.translation_keys.find({}, {'_id': 0}).to_list(1000)
    for t in translations:
        if isinstance(t.get('updated_at'), str):
            t['updated_at'] = datetime.fromisoformat(t['updated_at'])
    return translations

@api_router.post("/admin/translations", response_model=TranslationKey)
async def create_translation(data: TranslationKey, user: User = Depends(get_current_user)):
    if user.role != 'Admin_Directeur':
        raise HTTPException(status_code=403, detail="Accès réservé à la Direction commerciale")
    
    translation = TranslationKey(**data.model_dump(), updated_by=user.id)
    trans_dict = translation.model_dump()
    trans_dict['updated_at'] = trans_dict['updated_at'].isoformat()
    await db.translation_keys.insert_one(trans_dict)
    return translation

@api_router.put("/admin/translations/{key_id}", response_model=TranslationKey)
async def update_translation(key_id: str, data: dict, user: User = Depends(get_current_user)):
    if user.role != 'Admin_Directeur':
        raise HTTPException(status_code=403, detail="Accès réservé à la Direction commerciale")
    
    data['updated_by'] = user.id
    data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.translation_keys.update_one(
        {'id': key_id},
        {'$set': data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Clé de traduction non trouvée")
    
    updated = await db.translation_keys.find_one({'id': key_id}, {'_id': 0})
    if isinstance(updated.get('updated_at'), str):
        updated['updated_at'] = datetime.fromisoformat(updated['updated_at'])
    return TranslationKey(**updated)

@api_router.get("/admin/export-data")
async def export_all_data(user: User = Depends(get_current_user)):
    if user.role != 'Admin_Directeur':
        raise HTTPException(status_code=403, detail="Accès réservé à la Direction commerciale")
    
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Export Users
        ws_users = wb.create_sheet("Utilisateurs")
        ws_users.append(["ID", "Nom", "Email", "Rôle", "Division", "Région", "Date création"])
        users = await db.users.find({}, {'_id': 0, 'password_hash': 0}).to_list(1000)
        for u in users:
            ws_users.append([
                u.get('id', ''),
                u.get('name', ''),
                u.get('email', ''),
                u.get('role', ''),
                u.get('division', ''),
                u.get('region', ''),
                u.get('created_at', '')
            ])
        
        # Export Comptes
        ws_comptes = wb.create_sheet("Clients_Prospects")
        ws_comptes.append([
            "ID", "Raison Sociale", "Division", "Région", "Adresse", "Ville", "Code Postal",
            "Secteur", "Taille", "Contact Nom", "Contact Poste", "Contact Email", 
            "Contact Téléphone", "Source", "Créé par", "Date création"
        ])
        comptes = await db.comptes.find({}, {'_id': 0}).to_list(1000)
        for c in comptes:
            ws_comptes.append([
                c.get('id', ''),
                c.get('raison_sociale', ''),
                c.get('division', ''),
                c.get('region', ''),
                c.get('adresse', ''),
                c.get('ville', ''),
                c.get('code_postal', ''),
                c.get('secteur', ''),
                c.get('taille', ''),
                c.get('contact_nom', ''),
                c.get('contact_poste', ''),
                c.get('contact_email', ''),
                c.get('contact_telephone', ''),
                c.get('source', ''),
                c.get('created_by', ''),
                c.get('created_at', '')
            ])
        
        # Export Opportunités
        ws_opps = wb.create_sheet("Opportunités")
        ws_opps.append([
            "ID", "Compte ID", "Type Besoin", "Volumes Estimés", "Températures", "Fréquence",
            "Marchandises", "Départ", "Arrivée", "Contraintes Horaires", "Urgence",
            "Commercial Responsable", "Date Premier Contact", "Canal", "Statut",
            "Montant Estimé", "Prochaine Relance", "Commentaires", "Date création"
        ])
        opps = await db.opportunites.find({}, {'_id': 0}).to_list(1000)
        for o in opps:
            ws_opps.append([
                o.get('id', ''),
                o.get('compte_id', ''),
                o.get('type_besoin', ''),
                o.get('volumes_estimes', ''),
                o.get('temperatures', ''),
                o.get('frequence', ''),
                o.get('marchandises', ''),
                o.get('depart', ''),
                o.get('arrivee', ''),
                o.get('contraintes_horaires', ''),
                o.get('urgence', ''),
                o.get('commercial_responsable', ''),
                o.get('date_premier_contact', ''),
                o.get('canal', ''),
                o.get('statut', ''),
                o.get('montant_estime', ''),
                o.get('prochaine_relance', ''),
                o.get('commentaires', ''),
                o.get('created_at', '')
            ])
        
        # Export Quality Records
        ws_quality = wb.create_sheet("Fiches_Qualité")
        ws_quality.append([
            "ID", "Compte ID", "Division", "Région", "Période", "Type Prestation",
            "Taux Service", "Nb Incidents", "Score Satisfaction", "Commentaires", "Date création"
        ])
        quality_records = await db.quality_records.find({}, {'_id': 0}).to_list(1000)
        for q in quality_records:
            ws_quality.append([
                q.get('id', ''),
                q.get('compte_id', ''),
                q.get('division', ''),
                q.get('region', ''),
                q.get('periode', ''),
                q.get('type_prestation', ''),
                q.get('taux_service', ''),
                q.get('nb_incidents', ''),
                q.get('score_satisfaction', ''),
                q.get('commentaires', ''),
                q.get('created_at', '')
            ])
        
        # Export Incidents
        ws_incidents = wb.create_sheet("Incidents")
        ws_incidents.append([
            "ID", "Quality Record ID", "Type", "Gravité", "Description", "Statut",
            "Action Corrective", "Date Clôture", "Date création"
        ])
        incidents = await db.incidents.find({}, {'_id': 0}).to_list(1000)
        for i in incidents:
            ws_incidents.append([
                i.get('id', ''),
                i.get('quality_record_id', ''),
                i.get('type', ''),
                i.get('gravite', ''),
                i.get('description', ''),
                i.get('statut', ''),
                i.get('action_corrective', ''),
                i.get('closed_at', ''),
                i.get('created_at', '')
            ])
        
        # Export Survey Responses
        ws_surveys = wb.create_sheet("Enquêtes_Satisfaction")
        ws_surveys.append([
            "ID", "Compte ID", "Division", "Période", "Note Globale", 
            "Commentaires", "Date soumission"
        ])
        survey_responses = await db.survey_responses.find({}, {'_id': 0}).to_list(1000)
        for s in survey_responses:
            ws_surveys.append([
                s.get('id', ''),
                s.get('compte_id', ''),
                s.get('division', ''),
                s.get('periode', ''),
                s.get('note_globale', ''),
                s.get('commentaires', ''),
                s.get('submitted_at', '')
            ])
        
        # Style headers
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Return file
        return FileResponse(
            temp_file.name,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f'export_als_groupe_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error exporting data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export: {str(e)}")

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
